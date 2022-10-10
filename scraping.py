import openpyxl
from selenium import webdriver
from tkinter import *
from tkinter.ttk import *

import tkinter as tk
from tkinter import ttk
from tkinter import filedialog as fd
from tkinter.messagebox import showinfo


fileURL = ""

searchResult = []

def fileScraping():
    global fileURL
    
    if fileURL == "":
        return
    
    global searchResult
    searchResult = []
    
    # option = webdriver.ChromeOptions()
    # option.add_argument('headless')

    # browser = webdriver.Chrome(options=option)
    browser = webdriver.Chrome()

    # Define variable to load the dataframe
    dataframe = openpyxl.load_workbook(fileURL)  
    
    # Define variable to read sheet
    dataframe1 = dataframe.active
    rows = dataframe1.iter_rows(1, dataframe1.max_row)
    cnt = 0
    
    for row in rows:
        resultString = "\""
        for col in range(0, 2):
            strArr = row[col].value.split()
            for indexStr in strArr:
                resultString += indexStr + "+"
        if cnt != 0:
            resultUrl = "https://duckduckgo.com/?q=" + resultString +"&t=h_&ia=web"
            browser.get(resultUrl)

            results = browser.find_element("id","links")
            
            data = results.text.split("\n")
            row[2].value = data[0]
            row[3].value = data[1]
            
            temp = []
            for i in range(0, 4):
                temp.append(row[i].value)
            searchResult.append(temp)
        cnt += 1
        
    dataframe.save("output.xlsx")
    
    if len(searchResult) > 0:
        display()


def customScraping():
    if input_name.get() == "" and input_company.get() == "":
        return
    
    # option = webdriver.ChromeOptions()
    # option.add_argument('headless')

    # browser = webdriver.Chrome(options=option)
    browser = webdriver.Chrome()
    
    resultUrl = "https://duckduckgo.com/?q=" + input_name.get() + input_company.get() +"&t=h_&ia=web"
    browser.get(resultUrl)
    
    results = browser.find_element("id","links")
    
    data = results.text.split("\n")
    
    text_result.delete("1.0", 'end')
    text_result.insert(END, data[0] + "\n\n" + data[1] + "\n\n" + data[2])
    

# scraping function
def scraping():
    if combo_searchType.get() == 'file':
        fileScraping()
    if combo_searchType.get() == 'input':
        customScraping()


# select file dialog function
def select_file():
    filetypes = (
        ('All files', '*.*'),
        ('text files', '*.txt')
    )

    filename = fd.askopenfilename(
        title='Open a file',
        initialdir='/',
        filetypes=filetypes)

    global fileURL 
    fileURL = filename
    
    label_filePath['text'] = fileURL

    showinfo(
        title='Selected File',
        message=filename
    )



def searchTypeChange(event):
    input_name.delete(0, END)
    input_company.delete(0, END)
    text_result.delete("1.0", 'end')
    
    if combo_searchType.get() == "file":
        root.geometry('1200x370')    
        btn_open.place(x=50, y=80)       
        label_filePath.place(x=130, y=82)
        frame.place(x=20, y=120)
        
        label_searchProperty_name.place_forget()
        label_searchProperty_company.place_forget()
        input_name.place_forget()
        input_company.place_forget()
        text_result.place_forget()
        
        btn_run.place(x=500, y=80)
        btn_run.config(width=15)
        
    if combo_searchType.get() == 'input':
        global fileURL
        fileURL = ""
        label_filePath['text'] = ""
        display_table.delete(*display_table.get_children())
        
        root.geometry('500x500')
        btn_open.place_forget()
        label_filePath.place_forget()
        frame.place_forget()
        
        label_searchProperty_name.place(x=30, y=82)
        label_searchProperty_company.place(x=20, y=120)
        
        input_name.place(x=100, y=82)
        input_company.place(x=100, y=120)
        text_result.place(x=100, y=150)
        
        btn_run.place(x=50, y=430)    
        btn_run.config(width=68)  
        
    return


# Dialog
root = tk.Tk()
root.title('Web Scraping')
root.resizable(False, False)
root.geometry('1200x370')



label_searchType = ttk.Label(
    root,
    text='Searching Type :',
    width=15
)
label_searchType.pack(expand=True)
label_searchType.place(x=30, y=30)


# combo box
combo_searchType = ttk.Combobox(
    root,
    text='file',
    width=10,
    textvariable='file')
combo_searchType['values'] = ['file', 'input']
combo_searchType['state'] = 'readonly'
combo_searchType.pack(expand=True)
combo_searchType.place(x=125, y=30)
combo_searchType.set('file')

combo_searchType.bind('<<ComboboxSelected>>', searchTypeChange)


# open button
btn_open = ttk.Button(
    root,
    text='Open a File',
    command=select_file
)
btn_open.pack(expand=True)
btn_open.place(x=50, y=80)


# label filepath
label_filePath = ttk.Label(
    root,
    text='',
    width=35,
    font=14,
    background ="lightgrey"
)
label_filePath.pack(expand=True)
label_filePath.place(x=130, y=82)


# input name
input_name = ttk.Entry(root, width=60)
input_name.pack(expand=True)
input_name.place(x=100, y=82)
input_name.place_forget()


# input company
input_company = ttk.Entry(root, width=60)
input_company.pack(expand=True)
input_company.place(x=100, y=120)
input_company.place_forget()


# label name
label_searchProperty_name = ttk.Label(
    root,
    text='Name',
    width=8,
)
label_searchProperty_name.pack(expand=True)
label_searchProperty_name.pack_forget()


# label company
label_searchProperty_company = ttk.Label(
    root,
    text='Company',
    width=10,
)
label_searchProperty_company.pack(expand=True)
label_searchProperty_company.pack_forget()


# run button
btn_run = ttk.Button(
    root,
    text='Run Scraping',
    command=scraping,
    width=15
)
btn_run.pack()
btn_run.place(x=500, y=80)


text_result = Text(root, width=45, height=15)
text_result.pack()
text_result.place(x=100, y=150)
text_result.place_forget()


def display():
    for i in range(0, len(searchResult)):
        display_table.insert(parent='', index='end', iid=i, text='', 
            values=((i+1), searchResult[i][0], searchResult[i][1], searchResult[i][2], searchResult[i][3]))
        

frame = Frame(root)
frame.pack()
frame.place(x=20, y=120)

display_table = ttk.Treeview(frame)

display_table['columns'] = ('item_id', 'item_name', 'item_company', 'item_position', 'item_url')

display_table.column("#0", width=0, stretch=NO)
display_table.column("item_id",anchor=CENTER, width=30)
display_table.column("item_name",anchor=W,width=120)
display_table.column("item_company",anchor=W,width=120)
display_table.column("item_position",anchor=W,width=450)
display_table.column("item_url",anchor=W,width=450)

display_table.heading("#0",text="",anchor=CENTER)
display_table.heading("item_id",text="Id",anchor=CENTER)
display_table.heading("item_name",text="Name",anchor=CENTER)
display_table.heading("item_company",text="Company",anchor=CENTER)
display_table.heading("item_position",text="Position",anchor=CENTER)
display_table.heading("item_url",text="Url",anchor=CENTER)

display_table.pack()


# run the application
root.mainloop()